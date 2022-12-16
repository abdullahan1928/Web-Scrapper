from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active

# set column width


def set_column_widths():
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 20

# set column names


def set_column_names():
    ws.cell(row=1, column=1, value="S.No")
    ws.cell(row=1, column=2, value="Ghrana Number")
    ws.cell(row=1, column=3, value="CNIC")
    ws.cell(row=1, column=4, value="Name")
    ws.cell(row=1, column=5, value="Phone Number")


set_column_widths()
set_column_widths()
